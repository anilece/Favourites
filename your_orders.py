from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import selenium.common.exceptions
import openpyxl


#give the account details
username="priyann+nonvermont2@amazon.com"
password="123456"




path="/Users/anilrz/Documents/nonvermont1.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
col=sheet.max_column
def find_element(id):
    return driver.find_element_by_id(id)

def wait_exp(xpath):
    wait=WebDriverWait(driver,20)
    element=wait.until(EC.element_to_be_clickable((By.XPATH,xpath)))
    return element

driver=webdriver.Chrome(executable_path="/Users/anilrz/Downloads/chromedriver")
driver.get("https://www.amazon.com/ap/signin?_encoding=UTF8&openid.assoc_handle=usflex&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.mode=checkid_setup&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&openid.ns.pape=http%3A%2F%2Fspecs.openid.net%2Fextensions%2Fpape%2F1.0&openid.pape.max_auth_age=0&openid.return_to=https%3A%2F%2Fwww.amazon.com%2Fgp%2Fyourstore%2Fhome%3Fie%3DUTF8%26action%3Dsign-out%26path%3D%252Fgp%252Fyourstore%252Fhome%26ref_%3Dnav_youraccount_signout%26signIn%3D1%26useRedirectOnSuccess%3D1")

find_element(id="ap_email").send_keys(username)
find_element(id="continue").click()
find_element(id="ap_password").send_keys(password)
find_element(id="signInSubmit").click()
wait_exp('//*[@id="nav-orders"]/span[2]').click() #your orders
wait_exp('//*[@id="orderTypeMenuContainer"]/ul/li[5]').click() # digital orders
order_ids=[]
containers= driver.find_elements_by_xpath('//*[@id="ordersContainer"]/div')
back=len(containers)
orders_current_page=driver.find_elements_by_xpath("//span[@class='a-color-secondary value']")
order_name = []
try:
    while ((EC.element_to_be_clickable("//*[@id='ordersContainer']/div[@class='a-row']/div/ul/li[@class='a-last']"))):
        driver.implicitly_wait(10)

        orders_current_page=driver.find_elements_by_xpath("//span[@class='a-color-secondary value']")

        ordercost = driver.find_elements_by_xpath(
            '//*[@id="ordersContainer"]/div[@class="a-box-group a-spacing-base order"]/div[1]/div/div/div/div[1]/div/div[2]/div[2]/span')
        orders = driver.find_elements_by_class_name("a-fixed-left-grid-inner")
        ordername = []
        for i in range(len(orders)):
            p = (orders[i].text.split("\n"))
            if p[0] != "":
                ordername.append(p[0])
        c=1
        temp=[]
        w=0
        for i in range(len(orders_current_page)):

            if c<4:
                c+=1
                temp.append(orders_current_page[i].text)

                if c==3:
                    w+=1
                    if temp[1] != "$0.00":
                        order_ids.append(temp)
                        order_name.append(ordername[w-1])
                        print(temp)
            elif c>=4:
                c=1
                temp=[]
                c+=1
                temp.append(orders_current_page[i].text)

        if (EC.element_to_be_clickable("//*[@id='ordersContainer']/div[@class='a-row']/div/ul/li[@class='a-last']")):
                next_but=driver.find_element_by_xpath("//*[@id='ordersContainer']/div[@class='a-row']/div/ul/li[@class='a-last']")
                next_but.click()
except(selenium.common.exceptions.NoSuchElementException) or (selenium.common.exceptions.TimeoutException):
        pass
e=1
print(len(order_ids))

for i in range(0,len(order_ids)):
    for j in range(1,4):
        sheet.cell(row=i+1,column=j).value=order_ids[i][j-1]

for i in range(1,len(order_name)+1):
    sheet.cell(row=i,column=4).value=order_name[i-1]
workbook.save(path)

