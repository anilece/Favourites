from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import selenium.common.exceptions
import openpyxl
path="/Users/anilrz/Documents/nonvermont1.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
col=sheet.max_column
print(rows,col)
print(sheet.cell(row=3,column=1).value)
# for i in range(0,len(order_ids)):
#     for j in range(1,4):
#         sheet.cell(row=i+1,column=j).value=order_ids[i][j-1]
#         sheet.cell(row=e,column=4).value=order_name[e-1]
#         e+=1
# workbook.save(path)

